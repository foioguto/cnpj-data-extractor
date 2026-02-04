import openpyxl
from openpyxl.utils import column_index_from_string
from typing import Generator, Tuple, Dict, Any
from .config import config

class ExcelHandler:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.fields_map = config.fields_map
        try:
            # Loads the workbook into memory to allow editing
            self.wb = openpyxl.load_workbook(file_path)
            self.ws = self.wb.active
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {file_path}")

    def get_cnpjs_generator(self, target_column: str, start_row: int) -> Generator[Tuple[int, str], None, None]:
        """
        Yields a tuple: (row_number, cnpj_value).
        Acts as a 'Queue', delivering one item at a time for processing.
        """
        
        col_idx = column_index_from_string(target_column)

        # Iterates from start_row to the last filled row in that specific column
        for row in self.ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                # Returns the current row number and the cleaned string value
                yield cell.row, str(cell.value).strip()


    def update_row(self, row_number: int, data: Dict[str, Any], columns: list[int]):
            """
            Write data on row and column specified, S2
            """

            for index, field in enumerate(self.fields_map):
                if index < len(columns):
                    col_idx = columns[index] 
                    value = data.get(field, "")
                    
                    self.ws.cell(row=row_number, column=col_idx, value=value)


    def save_file(self):
        """Persists changes to disk."""
        print(f"💾 Saving Excel file to {self.file_path}...")
        self.wb.save(self.file_path)
        self.wb.close()
        